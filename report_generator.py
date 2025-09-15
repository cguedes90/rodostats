# -*- coding: utf-8 -*-
"""
Sistema de Relatórios Automatizados - Rodo Stats
Desenvolvido por InovaMente Labs
"""

import os
from datetime import datetime, timedelta
from io import BytesIO
import tempfile

# Relatórios PDF
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Relatórios Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference

class ReportGenerator:
    """Gerador de relatórios PDF e Excel para frotas"""

    def __init__(self, app_context=None):
        self.app_context = app_context

    def generate_fleet_report_pdf(self, fleet, fleet_stats, vehicles, fuel_records, period_days=30):
        """Gera relatório executivo da frota em PDF"""

        # Criar arquivo temporário
        buffer = BytesIO()

        # Configurar documento
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#2c3e50')
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#34495e')
        )

        normal_style = styles['Normal']

        # Conteúdo do relatório
        story = []

        # Cabeçalho
        title = Paragraph(f"🚛 RELATÓRIO EXECUTIVO DE FROTA", title_style)
        story.append(title)

        company_info = Paragraph(
            f"<b>Empresa:</b> {fleet.company_name}<br/>"
            f"<b>CNPJ:</b> {fleet.cnpj}<br/>"
            f"<b>Período:</b> {period_days} dias<br/>"
            f"<b>Gerado em:</b> {datetime.now().strftime('%d/%m/%Y às %H:%M')}<br/>",
            normal_style
        )
        story.append(company_info)
        story.append(Spacer(1, 20))

        # KPIs Principais
        kpis_heading = Paragraph("📊 INDICADORES PRINCIPAIS", heading_style)
        story.append(kpis_heading)

        kpis_data = [
            ['Métrica', 'Valor'],
            ['Total de Veículos', f"{fleet_stats['total_vehicles']}"],
            ['Total Gasto', f"R$ {fleet_stats['total_spent']:.2f}"],
            ['Total de Litros', f"{fleet_stats['total_liters']:.1f} L"],
            ['Consumo Médio da Frota', f"{fleet_stats['avg_consumption']:.1f} km/L"],
            ['Abastecimentos no Período', f"{fleet_stats['total_records_30d']}"],
        ]

        kpis_table = Table(kpis_data, colWidths=[3*inch, 2*inch])
        kpis_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))

        story.append(kpis_table)
        story.append(Spacer(1, 20))

        # Ranking de Veículos
        vehicles_heading = Paragraph("🚗 RANKING DE EFICIÊNCIA POR VEÍCULO", heading_style)
        story.append(vehicles_heading)

        # Calcular eficiência por veículo
        vehicle_efficiency = []
        for vehicle in vehicles:
            vehicle_records = [r for r in fuel_records if r.vehicle_id == vehicle.id]
            if len(vehicle_records) >= 2:
                total_km = sum(r.kilometers for r in vehicle_records if r.kilometers)
                total_liters = sum(r.liters for r in vehicle_records)
                avg_consumption = total_km / total_liters if total_liters > 0 else 0
                vehicle_efficiency.append((vehicle, avg_consumption, len(vehicle_records)))

        # Ordenar por eficiência
        vehicle_efficiency.sort(key=lambda x: x[1], reverse=True)

        vehicle_data = [['Posição', 'Veículo', 'Consumo (km/L)', 'Abastecimentos']]
        for i, (vehicle, consumption, records_count) in enumerate(vehicle_efficiency[:10], 1):
            vehicle_data.append([
                f"{i}º",
                f"{vehicle.brand} {vehicle.model} ({vehicle.license_plate})",
                f"{consumption:.1f}",
                f"{records_count}"
            ])

        if len(vehicle_data) > 1:
            vehicle_table = Table(vehicle_data, colWidths=[0.8*inch, 2.5*inch, 1.2*inch, 1.2*inch])
            vehicle_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            story.append(vehicle_table)
        else:
            story.append(Paragraph("Dados insuficientes para gerar ranking.", normal_style))

        story.append(Spacer(1, 20))

        # Recomendações
        recommendations_heading = Paragraph("💡 RECOMENDAÇÕES E INSIGHTS", heading_style)
        story.append(recommendations_heading)

        recommendations = []

        if fleet_stats['avg_consumption'] < 8:
            recommendations.append("⚠️ Consumo médio da frota abaixo do esperado. Verificar manutenções pendentes.")
        elif fleet_stats['avg_consumption'] > 12:
            recommendations.append("✅ Excelente eficiência da frota! Manter práticas atuais.")
        else:
            recommendations.append("📈 Consumo dentro da média. Há potencial para melhoria.")

        if fleet_stats['total_records_30d'] < len(vehicles) * 4:
            recommendations.append("📝 Baixa frequência de abastecimentos registrados. Incentivar uso do sistema.")

        recommendations.append("🔧 Implementar manutenção preventiva baseada em quilometragem.")
        recommendations.append("📊 Acompanhar custos semanalmente para detectar anomalias.")

        for rec in recommendations:
            story.append(Paragraph(f"• {rec}", normal_style))
            story.append(Spacer(1, 6))

        # Rodapé
        story.append(Spacer(1, 30))
        footer = Paragraph(
            "Relatório gerado automaticamente pelo Rodo Stats | InovaMente Labs",
            ParagraphStyle('Footer', parent=normal_style, fontSize=8, alignment=TA_CENTER, textColor=colors.grey)
        )
        story.append(footer)

        # Gerar PDF
        doc.build(story)

        # Retornar bytes do PDF
        pdf_data = buffer.getvalue()
        buffer.close()

        return pdf_data

    def generate_fleet_report_excel(self, fleet, fleet_stats, vehicles, fuel_records, period_days=30):
        """Gera relatório da frota em Excel com múltiplas abas"""

        # Criar workbook
        wb = Workbook()

        # === ABA 1: RESUMO EXECUTIVO ===
        ws_summary = wb.active
        ws_summary.title = "Resumo Executivo"

        # Cabeçalho
        ws_summary['A1'] = f"RELATÓRIO EXECUTIVO - {fleet.company_name}"
        ws_summary['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws_summary['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws_summary['A1'].alignment = Alignment(horizontal="center")
        ws_summary.merge_cells('A1:D1')

        # Informações da empresa
        ws_summary['A3'] = "Empresa:"
        ws_summary['B3'] = fleet.company_name
        ws_summary['A4'] = "CNPJ:"
        ws_summary['B4'] = fleet.cnpj
        ws_summary['A5'] = "Período:"
        ws_summary['B5'] = f"{period_days} dias"
        ws_summary['A6'] = "Gerado em:"
        ws_summary['B6'] = datetime.now().strftime('%d/%m/%Y %H:%M')

        # KPIs
        ws_summary['A8'] = "INDICADORES PRINCIPAIS"
        ws_summary['A8'].font = Font(size=14, bold=True, color="FFFFFF")
        ws_summary['A8'].fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        ws_summary.merge_cells('A8:B8')

        kpis = [
            ("Total de Veículos", fleet_stats['total_vehicles']),
            ("Total Gasto", f"R$ {fleet_stats['total_spent']:.2f}"),
            ("Total de Litros", f"{fleet_stats['total_liters']:.1f} L"),
            ("Consumo Médio", f"{fleet_stats['avg_consumption']:.1f} km/L"),
            ("Abastecimentos", fleet_stats['total_records_30d'])
        ]

        for i, (metric, value) in enumerate(kpis, 9):
            ws_summary[f'A{i}'] = metric
            ws_summary[f'B{i}'] = value
            ws_summary[f'A{i}'].font = Font(bold=True)

        # === ABA 2: DETALHES POR VEÍCULO ===
        ws_vehicles = wb.create_sheet("Detalhes por Veículo")

        headers = ["Veículo", "Placa", "Consumo Médio", "Total Gasto", "Total Litros", "Abastecimentos"]
        for col, header in enumerate(headers, 1):
            cell = ws_vehicles.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Dados por veículo
        for row, vehicle in enumerate(vehicles, 2):
            vehicle_records = [r for r in fuel_records if r.vehicle_id == vehicle.id]

            total_km = sum(r.kilometers for r in vehicle_records if r.kilometers)
            total_liters = sum(r.liters for r in vehicle_records)
            total_cost = sum(r.total_cost for r in vehicle_records)
            avg_consumption = total_km / total_liters if total_liters > 0 else 0

            ws_vehicles.cell(row=row, column=1, value=f"{vehicle.brand} {vehicle.model}")
            ws_vehicles.cell(row=row, column=2, value=vehicle.license_plate)
            ws_vehicles.cell(row=row, column=3, value=f"{avg_consumption:.1f}")
            ws_vehicles.cell(row=row, column=4, value=total_cost)
            ws_vehicles.cell(row=row, column=5, value=total_liters)
            ws_vehicles.cell(row=row, column=6, value=len(vehicle_records))

        # === ABA 3: HISTÓRICO DE ABASTECIMENTOS ===
        ws_fuel = wb.create_sheet("Histórico Abastecimentos")

        fuel_headers = ["Data", "Veículo", "Placa", "Litros", "Valor", "Km", "Consumo", "Posto"]
        for col, header in enumerate(fuel_headers, 1):
            cell = ws_fuel.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Ordenar por data
        sorted_records = sorted(fuel_records, key=lambda x: x.date, reverse=True)

        for row, record in enumerate(sorted_records, 2):
            vehicle = next((v for v in vehicles if v.id == record.vehicle_id), None)

            ws_fuel.cell(row=row, column=1, value=record.date.strftime('%d/%m/%Y'))
            ws_fuel.cell(row=row, column=2, value=f"{vehicle.brand} {vehicle.model}" if vehicle else "N/A")
            ws_fuel.cell(row=row, column=3, value=vehicle.license_plate if vehicle else "N/A")
            ws_fuel.cell(row=row, column=4, value=record.liters)
            ws_fuel.cell(row=row, column=5, value=record.total_cost)
            ws_fuel.cell(row=row, column=6, value=record.kilometers or 0)
            ws_fuel.cell(row=row, column=7, value=f"{record.consumption:.1f}" if record.consumption else "N/A")
            ws_fuel.cell(row=row, column=8, value=record.gas_station or "N/A")

        # Ajustar largura das colunas
        for ws in [ws_summary, ws_vehicles, ws_fuel]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Salvar em buffer
        buffer = BytesIO()
        wb.save(buffer)
        excel_data = buffer.getvalue()
        buffer.close()

        return excel_data

    def save_report_to_file(self, report_data, filename, report_type="pdf"):
        """Salva relatório em arquivo"""

        reports_dir = "static/reports"
        if not os.path.exists(reports_dir):
            os.makedirs(reports_dir)

        file_path = os.path.join(reports_dir, filename)

        with open(file_path, 'wb') as f:
            f.write(report_data)

        return file_path

def generate_fleet_reports(fleet, period_days=30):
    """Função utilitária para gerar todos os relatórios de uma frota"""
    from app import db, Vehicle, FuelRecord

    # Buscar dados
    vehicles = Vehicle.query.filter_by(fleet_id=fleet.id, is_active=True).all()

    # Período de análise
    end_date = datetime.now()
    start_date = end_date - timedelta(days=period_days)

    fuel_records = db.session.query(FuelRecord).join(Vehicle).filter(
        Vehicle.fleet_id == fleet.id,
        FuelRecord.date >= start_date,
        FuelRecord.date <= end_date
    ).all()

    # Calcular estatísticas
    fleet_stats = {
        'total_vehicles': len(vehicles),
        'total_spent': sum(r.total_cost for r in fuel_records),
        'total_liters': sum(r.liters for r in fuel_records),
        'total_records_30d': len(fuel_records)
    }

    # Calcular consumo médio da frota
    consumptions = []
    for vehicle in vehicles:
        vehicle_records = [r for r in fuel_records if r.vehicle_id == vehicle.id]
        if len(vehicle_records) >= 2:
            total_km = sum(r.kilometers for r in vehicle_records if r.kilometers)
            total_liters = sum(r.liters for r in vehicle_records)
            if total_liters > 0:
                consumptions.append(total_km / total_liters)

    fleet_stats['avg_consumption'] = sum(consumptions) / len(consumptions) if consumptions else 0

    # Gerar relatórios
    generator = ReportGenerator()

    pdf_data = generator.generate_fleet_report_pdf(
        fleet, fleet_stats, vehicles, fuel_records, period_days
    )

    excel_data = generator.generate_fleet_report_excel(
        fleet, fleet_stats, vehicles, fuel_records, period_days
    )

    return pdf_data, excel_data, fleet_stats